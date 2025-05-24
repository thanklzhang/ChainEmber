# -*- coding: utf-8 -*-
#!/usr/bin/python

# import copy
# import getopt
import json
import os.path
# import re
import sys
import codecs
import re 

from jinja2 import Environment, FileSystemLoader

from openpyxl import load_workbook

from enum import Enum

class OpType(Enum):
    cs = 1
    json = 2

all_json_data_list = []

def main(argv):
    print("")
def gen(args):
    input = args.get('in_path')
    out = args.get('out_path')
    opType = args.get('op_type')
    resOutPath = args.get('res_out_path')
    battle_config_cs_path = args.get('battle_config_cs_path')
    battle_config_impl_cs_path = args.get('battle_config_impl_cs_path')
    if OpType.cs == opType:
        #gen cs
        list_dirs = os.walk(input)
        print("start generate cs ... : from : " + input)
        for root,dirs,files in list_dirs:
            for f in files:
                splitStr = os.path.splitext(f)
                path_without_ext = splitStr[0]
                ext = splitStr[1]
                if ext == '.xlsx' and not '~$' in path_without_ext:
                    input_path = os.path.join(root, f)
                    # cs define files
                    gen_cs_define_file(input_path,out,battle_config_cs_path,battle_config_impl_cs_path)

        print("finish cs finish! : to : " + out)
    elif OpType.json == opType:
        #gen json
        print("start generate json ... : from : " + input)
        tableDataPathList = []
        list_dirs2 = os.walk(input)
        for root,dirs,files in list_dirs2:
            for f in files:
                splitStr = os.path.splitext(f)
                path_without_ext = splitStr[0]
                ext = splitStr[1]
                if ext == '.xlsx' and not '~$' in path_without_ext:
                    input_path = os.path.join(root, f)
                    json_data,class_name = gen_json_file(input_path,out)
                    tableDataPathList.append(class_name + '.json');
                    if class_name == 'ResourceConfig':
                        gen_res_id_dic(json_data,resOutPath)
                    # if class_name == 'BattleTrigger':
                       # gen_battle_trigger_dic(json_data,battleTriggerOutPath)
        # gen_table_path_file(tableDataPathList,tableDataOutPath)
        print("finish json! : to : " + out)
      
        
     
        
        
    


################## 生成单个cs 定义文件
def gen_cs_define_file(input_file,output_dictionary,battle_config_cs_path,battle_config_impl_cs_path):
    data_list,class_name = get_table_head_list(input_file)
    
    #过滤掉转表符的数据 和 id 项(因为 BaseTable 中已经有 id 的定义了)
    #data_list = filter(lambda x: x.name != '#', data_list)#导致在 jiaja 中有些问题 不好使 待测
    for i in range(len(data_list)-1, -1, -1):
        value = data_list[i]
        if value.name == '#' or value.name == 'id':
            data_list.remove(value)

    env = Environment(loader=FileSystemLoader('./'))

    #通用配置 cs 定义
    template = env.get_template('template/cs_template_class_define.j2')
    result = template.render( name=class_name,list=data_list)

    if not os.path.exists(output_dictionary):
        #os.mkdir(output_dictionary)
        os.makedirs(output_dictionary, exist_ok=True)

    with codecs.open(f'{output_dictionary}/{class_name}.cs', "w", 'utf8') as f:
        f.write(result)
        
    #战斗配置定义cs
    if battle_config_cs_path != None:
        template = env.get_template('template/cs_template_class_define_battle_config.j2')
        result = template.render( name=class_name,list=data_list)
    
        if not os.path.exists(battle_config_cs_path):
            os.makedirs(battle_config_cs_path, exist_ok=True)
    
        with codecs.open(f'{battle_config_cs_path}/I{ class_name}.cs', "w", 'utf8') as f:
            f.write(result)

    #战斗配置的实现的定义cs
    if battle_config_impl_cs_path != None:
        template = env.get_template('template/cs_template_class_define_battle_config_client_impl.j2')
        result = template.render( name=class_name,list=data_list)

        if not os.path.exists(battle_config_impl_cs_path):
            os.makedirs(battle_config_impl_cs_path, exist_ok=True)

        with codecs.open(f'{battle_config_impl_cs_path}/{ class_name}_Impl.cs', "w", 'utf8') as f:
            f.write(result)
        
    print("finish generate cs_define : " + class_name + ".cs")

################## 生成单个json 文件
def gen_json_file(input_file,output_dictionary):
    json_data,class_name = get_table_data_list(input_file,output_dictionary)
    json_str = json.dumps(json_data,sort_keys=False, indent=4, separators=(',', ': '))
    
    with codecs.open(f'{output_dictionary}/{class_name}.json', "w", 'utf8') as f:
        f.write(json_str)
    print("finish generate json : " + class_name + ".json")
    return json_data,class_name
    
    
################## 生成资源 id 对应表
def gen_res_id_dic(json_data,output_dictionary):
    _list = []
    
    
    for value in json_data:
        _id = value['id']
        name = value['name']
        path = value['path']
        ext = value['ext']
        
        fullPath = '' + path + '/' + name + "." + ext
        obj = {}
        obj['name'] = name
        obj['id'] = _id
        obj['fullPath'] = fullPath
        _list.append(obj)
        
    
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template_res_dic_temp.j2')
    result = template.render(list=_list)
    
    with codecs.open(f'{output_dictionary}/ResDefine.cs', "w", 'utf8') as f:
        f.write(result)
        
    print("finish res dic cs")

################## 生成战斗触发器路径对应表
def gen_battle_trigger_dic(json_data,output_dictionary):
    
    _dic = {}
    for value in json_data:
        path = value['scriptPath']
       
        if not _dic.__contains__(path):
            rootPath = 'CommonData'
            fullPath = os.path.join(rootPath,path)
            _list = []
          
            list_dirs2 = os.walk(fullPath)
            for root,dirs,files in list_dirs2:
                for f in files:
                    currFullPath = os.path.join(root,f)
                    splitStr = os.path.splitext(f)
                    path_without_ext = splitStr[0]
                    ext = splitStr[1]
                    if ext == '.json':
                        localPath = currFullPath.split(rootPath)[1][1:].replace('\\','/')
                        _list.append(localPath)
                        
            
            _dic[path] = _list
            
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_battle_trigger_dic_temp.j2')
    result = template.render(dic=_dic)
    
    with codecs.open(f'{output_dictionary}/BattleTriggerPathDefine.cs', "w", 'utf8') as f:
        f.write(result)
        
    print("finish res dic cs")

def gen_table_path_file(pathList,output_dictionary):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_table_path_temp.j2')
    result = template.render(list=pathList)
    
    with codecs.open(f'{output_dictionary}/TablePathDefine.cs', "w", 'utf8') as f:
        f.write(result)

################## 生成单个读取 json 的读取器(.cs)
def gen_load_json_file(input_file,output_dictionary):
    data_list,class_name = get_table_head_list(input_file)
    
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template_load_json.j2')
    result = template.render( name=class_name)

    if not os.path.exists(output_dictionary):
        os.mkdir(output_dictionary)

    with codecs.open(f'{output_dictionary}/{class_name}Store.cs', "w", 'utf8') as f:
        f.write(result)
    print("finish generate cs_load_json : " + class_name + "Store.cs")

################## 获取表格的前几行表头数据
def get_table_head_list(input_file):
    # print(input_file)
    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    name_obj_list = ws[1]
    type_obj_list = ws[2]
    comment_obj_list = ws[3]
    #print('get_table_head_list : input_file : ' + input_file)
    head_list = []
    for col_index,name_obj in enumerate(name_obj_list):
        # #过滤转表符
        # if 0 == col_index:
        #     continue
       
        name_value = name_obj.value
        type_value = type_obj_list[col_index].value
        comment_value = comment_obj_list[col_index].value
        #print('sss : ' + name_value + ' ' + type_value + ' ' + comment_value)
        head_data = HeadDefine(name_value,type_value,comment_value)
        head_list.append(head_data)
        # print(head_data.name + " " + head_data.type + " " + head_data.comment)

    #print(input_file)
    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]

    return head_list,class_name
   

################## 获取表格的数据部分(json)
def get_table_data_list(input_file, output_dictionary):
    # print(input_file)
    head_list,class_name = get_table_head_list(input_file)

    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    json_data_list = []
    for row_index,row_data in enumerate(ws):
        curr_data = {}
        if row_index >= 3:
            is_gen_data = True
            for col_index,data_obj in enumerate(row_data):
                #转表符 不生成数据
                if 0 == col_index:
                    if data_obj.value != '#':
                        is_gen_data = False
                        break
                
                field_name = head_list[col_index].name
                field_type = head_list[col_index].type
                if field_name == '#':
                    continue
                if data_obj.value == None:
                    if field_type == 'string':
                        curr_data[field_name] = ""
                    if field_type == 'int':
                        curr_data[field_name] = 0
                    if field_type == 'bool':
                        curr_data[field_name] = ""
                    if 'List' in field_type:
                        curr_data[field_name] = []
                else:
                    #base
                    if field_type == 'string':
                        curr_data[field_name] = str(data_obj.value)
                    if field_type == 'int':
                        curr_data[field_name] = int(data_obj.value)
                    if field_type == 'bool':
                        curr_data[field_name] = str(data_obj.value)
                        
                    #List<>
                    if field_type == 'List<int>':
                        valueStr = str(data_obj.value)
                        curr_data[field_name] = [int(item) for item in valueStr.split(',')]
                    if field_type == 'List<bool>':
                        valueStr = str(data_obj.value)
                        curr_data[field_name] = [str(item) for item in valueStr.split(',')]
                    if field_type == 'List<string>':
                        valueStr = str(data_obj.value)
                        curr_data[field_name] = [str(item) for item in valueStr.split(',')]
                    
                    #List<List<>>
                    if 'List<List<' in field_type:
                        parseType = extract_type_from_list_pattern(field_type)
                        
                        if parseType == None:
                            continue
                            
                        valueStr = str(data_obj.value)
                        result = []
                        outArr = [str(item) for item in valueStr.split('|')]
                        
                        if parseType == 'int':
                            for inner in outArr:
                                result.append([int(item) for item in inner.split(',')])
                        else:
                            for inner in outArr:
                                result.append([str(item) for item in inner.split(',')])
                            
                        curr_data[field_name] = result
                        
            if is_gen_data:
                json_data_list.append(curr_data)

    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]
    return json_data_list,class_name
   
 
  
def extract_type_from_list_pattern(s):  
    # 正则表达式匹配 List<List<type>> 模式  
    match = re.match(r'List<List<([^>]+)>>', s)  
    if match:  
        # 如果匹配成功，返回类型  
        return match.group(1)  
    else:  
        # 如果匹配失败，返回 None  
        return None  

def capitalize(string, lower_rest=False):
    ''' 字符转换
    :param string: 传入原始字符串
    :param lower_rest: bool, 控制参数--是否将剩余字母都变为小写
    :return: 改变后的字符
    '''
    
    return string[:1].upper() + (string[1:].lower() if lower_rest else string[1:])

##################
class HeadDefine():
    def __init__(self,name,type,comment):
        self.name = name
        self.type = type
        self.comment = comment
        self.up_name = capitalize(self.name)


##################
if __name__ == "__main__":
    main(sys.argv)
