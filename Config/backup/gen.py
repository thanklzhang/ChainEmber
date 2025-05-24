# -*- coding: utf-8 -*-
#!/usr/bin/python

# import copy
# import getopt
import json
import os.path
# import re
import sys
import codecs

from jinja2 import Environment, FileSystemLoader

from openpyxl import load_workbook

################ 配置
# table 文件输入目录
table_input_dir = 'data_table'
# 客户端 cs define 文件输出目录
client_cs_out_dir = "../001_GameFramework_Client/Assets/Script/Data/Define/Table"
# 客户端 json 文件输出目录
client_json_out_dir = "../001_GameFramework_Client/Assets/BuildRes/TableData"
# 客户端 cs load_json 文件输出目录
#client_cs_load_json_out_dir = "../001_GameFramework_Client/Assets/Script/Manager/TableManager/TableStore"
#################

def main(argv):
   
    list_dirs = os.walk(table_input_dir)
    print("start generate cs ...")
    for root,dirs,files in list_dirs:
        for f in files:
            splitStr = os.path.splitext(f)
            path_without_ext = splitStr[0]
            ext = splitStr[1]
            if ext == '.xlsx' and not '~$' in path_without_ext:
                input_path = os.path.join(root, f)
                # cs define files
               
                gen_cs_define_file(input_path,client_cs_out_dir)
               
    print("finish all cs!")

    print("start generate json ...")
    list_dirs2 = os.walk(table_input_dir)
    for root,dirs,files in list_dirs2:
        for f in files:
            splitStr = os.path.splitext(f)
            path_without_ext = splitStr[0]
            ext = splitStr[1]
            if ext == '.xlsx' and not '~$' in path_without_ext:
                input_path = os.path.join(root, f)
                gen_json_file(input_path,client_json_out_dir)
                #gen_load_json_file(input_path,client_cs_load_json_out_dir)
                
    print("finish all json!")
    print("finish all !")
                


################## 生成单个cs 定义文件
def gen_cs_define_file(input_file,output_dictionary):
    data_list,class_name = get_table_head_list(input_file)
    
    #过滤掉转表符的数据 和 id 项(因为 BaseTable 中已经有 id 的定义了)
    #data_list = filter(lambda x: x.name != '#', data_list)#导致在 jiaja 中有些问题 不好使 待测
    for i in range(len(data_list)-1, -1, -1):
        value = data_list[i]
        if value.name == '#' or value.name == 'id':
            data_list.remove(value)

    # for value in data_list:
    #     print(value.name)
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template_class_define.j2')
    result = template.render( name=class_name,list=data_list)

    if not os.path.exists(output_dictionary):
        os.mkdir(output_dictionary)

    with codecs.open(f'{output_dictionary}/{class_name}.cs', "w", 'utf8') as f:
        f.write(result)
    print("finish generate cs_define : " + class_name + ".cs")

################## 生成单个json 文件
def gen_json_file(input_file,output_dictionary):
    json_data,class_name = get_table_data_list(input_file,output_dictionary)
    json_str = json.dumps(json_data,sort_keys=False, indent=4, separators=(',', ': '))
    
    with codecs.open(f'{output_dictionary}/{class_name}.json', "w", 'utf8') as f:
        f.write(json_str)

    print("finish generate json : " + class_name + ".json")

################## 生成单个读取 json 的读取器(.cs)
def gen_load_json_file(input_file,output_dictionary):
    data_list,class_name = get_table_head_list(input_file)
    
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template_load_json.j2')
    result = template.render( name=class_name)

    if not os.path.exists(output_dictionary):
        os.mkdir(output_dictionary)

    with codecs.open(f'{output_dictionary}/{class_name}Store.cs', "w", 'utf8') as f:
        f.write(result)
    print("finish generate cs_load_json : " + class_name + "Store.cs")

################## 获取表格的前几行表头数据
def get_table_head_list(input_file):
    # print(input_file)
    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    name_obj_list = ws[1]
    type_obj_list = ws[2]
    comment_obj_list = ws[3]
    #print('get_table_head_list : input_file : ' + input_file)
    head_list = []
    for col_index,name_obj in enumerate(name_obj_list):
        # #过滤转表符
        # if 0 == col_index:
        #     continue
       
        name_value = name_obj.value
        type_value = type_obj_list[col_index].value
        comment_value = comment_obj_list[col_index].value
        #print('sss : ' + name_value + ' ' + type_value + ' ' + comment_value)
        head_data = HeadDefine(name_value,type_value,comment_value)
        head_list.append(head_data)
        # print(head_data.name + " " + head_data.type + " " + head_data.comment)

    #print(input_file)
    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]

    return head_list,class_name
   

################## 获取表格的数据部分(json)
def get_table_data_list(input_file, output_dictionary):
    # print(input_file)
    head_list,class_name = get_table_head_list(input_file)

    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    json_data_list = []
    for row_index,row_data in enumerate(ws):
        curr_data = {}
        if row_index >= 3:
            is_gen_data = True
            for col_index,data_obj in enumerate(row_data):
                #转表符 不生成数据
                if 0 == col_index:
                    if data_obj.value != '#':
                        is_gen_data = False
                        break
                
                field_name = head_list[col_index].name
                field_type = head_list[col_index].type
                if field_name == '#':
                    continue
                if data_obj.value == None:
                    if field_type == 'string':
                        curr_data[field_name] = ""
                    if field_type == 'int':
                        curr_data[field_name] = 0
                else:
                    if field_type == 'string':
                        curr_data[field_name] = '"' + str(data_obj.value) + '"'
                    else:
                        curr_data[field_name] = data_obj.value
            if is_gen_data:
                json_data_list.append(curr_data)

    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]
    return json_data_list,class_name
   


def capitalize(string, lower_rest=False):
    ''' 字符转换
    :param string: 传入原始字符串
    :param lower_rest: bool, 控制参数--是否将剩余字母都变为小写
    :return: 改变后的字符
    '''
    
    return string[:1].upper() + (string[1:].lower() if lower_rest else string[1:])

##################
class HeadDefine():
    def __init__(self,name,type,comment):
        self.name = name
        self.type = type
        self.comment = comment
        self.up_name = capitalize(self.name)


##################
if __name__ == "__main__":
    main(sys.argv)
