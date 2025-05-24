# -*- coding: utf-8 -*-
#!/usr/bin/python

import copy
import getopt
import json
import os.path
import re
import sys
import codecs

from jinja2 import Environment, FileSystemLoader
# import openpyxl
from openpyxl import load_workbook

# Get Command Line Arguments

enums = []

def main(argv):
    input_file = ''
    output_dictionary = ''
    gen_type = ''

    try:
        opts, args = getopt.getopt(
            argv, "hi:o:f:", ["ifile=", "odir=", "gen="])
    except getopt.GetoptError:
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            sys.exit()
        elif opt in ("-i", "--ifile"):
            input_file = arg
        elif opt in ("-o", "--odir"):
            output_dictionary = arg
        elif opt in ('-g', '--gen'):
            gen_type = arg

    gen(input_file, output_dictionary, gen_type)

# create enums


def gen_enum(input_file, output_dictionary):
    wb = load_workbook(input_file)
    ws = wb.worksheets[0]

    # paser header

    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]
    class_name = class_name[0].upper() + class_name[1:]


    row_index = 0
    for row in ws.rows:
        row_index = row_index + 1
        if row_index < 4:
            continue

        if row[0].value is None:
            continue

        enumType = row[1].value
        name = row[2].value
        value = int(row[3].value)
        text = row[4].value
        findEnum = None
        for enum in enums:
            if enum["enumType"] == enumType:
                findEnum = enum
                break

        if findEnum is None:
            findEnum = {"enumType": enumType, "items": []}
            enums.append(findEnum)

        findEnum["items"].append({
            "name": name,
            "value": value,
            "text": text
        })

    enum_code = enum_data_to_cs_code(enums)
    with codecs.open(f'{output_dictionary}/EnumCode.cs', "w", 'utf8') as f:
        f.write(enum_code)


def gen(input_file, output_dictionary, gen_type,res_file_name,res_ids_out_path):
    # print(input_file)
    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    header = ws[1]
    data_type = ws[2]
    comment = ws[3]

    # paser header
    root = paser_header(header, data_type, comment)

    # generate tree
    # print root.find_child_by_path('event_1_param')
    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]
    class_name = class_name[0].upper() + class_name[1:]
    all = Node(class_name)
    all.type = 'array'
    # print(len(all.children))
    table_define = []
    row_index = 0
    for row in ws.rows:
        row_index = row_index + 1
        if row_index < 4:
            continue

        if row[0].value is None:
            continue

        row_root = copy.deepcopy(root)
        row_root.name = row_index - 3
        row_root.set_parent(all)

        col_index = 0
        
        resName = ""
        resId = 0
        resType = ""
        for cell in row:
            # print header[col_index].value
            # print dataType[col_index].value
            # print cell.value
            headerName = header[col_index].value
            child_node = row_root.find_child_by_path(headerName)
            # child_node.type_name = data_type[col_index].value
            child_node.value = convert_data_type(
                child_node.type_name, cell.value)
            
            #collect res path ids
            if class_name == res_file_name:
                if 'id' == headerName:
                    resId = cell.value
                if 'name'== headerName:
                    resName = cell.value
                if 'type' == headerName:
                    resType = cell.value
                                
            
            col_index = col_index + 1
            
        if class_name == res_file_name:
            if "ui" == resType:# 目前只生成 UI 预设资源的枚举 根据需求再改
                tempResKV = {"name":resName,"id":resId}
                table_define.append(tempResKV);
        # print(len(all.children))
        
    if class_name == res_file_name:
        #resPath = '../JekoClient/GameClient/JekoClient/Assets\Script\Data\TableData'
        gen_table_define_cs_code(res_ids_out_path,table_define);
        
    if not os.path.exists(output_dictionary):
        os.mkdir(output_dictionary)

    if gen_type == 'json':
        # generate json file
        obj = {}
        tree_to_object(all, obj)
        json_str = json.dumps(obj[class_name+'List'])
        # print(json_str)
        with codecs.open(f'{output_dictionary}/{class_name}.json', "w", 'utf8') as f:
            f.write(json_str)
    elif gen_type == 'cs':
        # generate csharp code
        cs_root = copy.deepcopy(root)
        # change type_name
        change_cs_type(cs_root)
        code = tree_to_cs_code(cs_root, class_name)
        with codecs.open(f'{output_dictionary}/{class_name}.cs', "w", 'utf8') as f:
            f.write(code)
     

            
    elif gen_type == 'go':
        # generate csharp code
        cs_root = copy.deepcopy(root)
        # change type_name
        change_cs_type(cs_root)
        code = tree_to_go_code(cs_root, class_name)
        with codecs.open(f'{output_dictionary}/{class_name}.go', "w", 'utf8') as f:
            f.write(code)


def change_cs_type(node):
    if node.type_name == 'float':
        node.type_name = 'float'

    for n in node.children:
        change_cs_type(n)


def convert_data_type(data_type, data):
    # print(data)
    if data:
        if data_type == 'int':
            return int(data)
        elif data_type == 'string':
            return str(data)
        elif data_type == 'float':
            return float(data)
        elif data_type == 'bool':
            return str(data) == 'True'
        else:
            return int(data)
    else:
        if data_type == 'int':
            return 0
        elif data_type == 'string':
            return ''
        elif data_type == 'float':
            return 0
        elif data_type == 'bool':
            return False
        else:
            return 0


class Node(object):
    def __init__(self, name):
        self.name = name
        self.parent = None
        self.children = []
        self.type = ''
        self.type_name = ''

    def set_parent(self, parent):
        self.parent = parent
        parent.children.append(self)

    def add_child(self, child):
        self.children.append(child)
        child.parent = self

    def find_child(self, child_name):
        for child in self.children:
            if child.name == child_name:
                return child
        return None

    def find_child_by_path(self, path):
        arr = path.split('_')
        parent = self
        for name in arr:
            find_node = parent.find_child(name)
            if find_node is None:
                return None
            else:
                parent = find_node

        return find_node


def paser_header(header, data_type, comment):
    root = Node('root')
    for col_index, field_name in enumerate(header):
        arr = field_name.value.split('_')
        parent = root
        for i, n in enumerate(arr):
            find_node = parent.find_child(n)
            if find_node is None:
                new_node = Node(n)

                if i < len(arr) - 1:
                    new_node.type_name = n[0].upper() + n[1:]
                    new_node.comment = ''
                else:
                    new_node.type_name = data_type[col_index].value
                    new_node.comment = comment[col_index].value

                if re.search(r'^\d+', n):
                    # print(n)
                    parent.type = 'array'
                    if i < len(arr) - 1:
                        new_node.type_name = parent.name[0].upper(
                        ) + parent.name[1:]
                    else:
                        new_node.type_name = data_type[col_index].value

                else:
                    parent.type = 'object'

                new_node.type = 'field'
                new_node.set_parent(parent)
                parent = new_node
                # print n
            else:
                parent = find_node
    return root


def tree_to_object(node, obj):
    # print(node.name)

    if node.type == 'object':
        cur_obj = {}
    elif node.type == 'array':
        cur_obj = []
    elif node.type == 'field':
        cur_obj = node.value
    else:
        cur_obj = {}

    if node.parent and node.parent.type == 'array':
        obj.append(cur_obj)
    else:
        if node.type == 'array':
            obj[node.name+'List'] = cur_obj
        else:
            obj[node.name] = cur_obj

    for child in node.children:
        tree_to_object(child, cur_obj)


def tree_to_cs_code(node, name):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template.j2')

    
    return template.render(node=node, name=name)

def tree_to_go_code(node, name):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/go_template.j2')
    return template.render(node=node, name=name)


def enum_data_to_cs_code(enums):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_enum.j2')
    return template.render(enums=enums)

def gen_auto_config_loader_cs_code(output_dictionary,names):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_auto_config_loader.j2')
    auto_config_load = template.render(names=names)
    with codecs.open(f'{output_dictionary}/AutoConfigDataLoader.cs', "w", 'utf8') as f:
        f.write(auto_config_load)
        
def gen_table_define_cs_code(output_dictionary,res_id_map):
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_table_define.j2')
    auto_config_load = template.render(res_id_map=res_id_map)
    with codecs.open(f'{output_dictionary}/TableDefine.cs', "w", 'utf8') as f:
        f.write(auto_config_load)


if __name__ == "__main__":
    main(sys.argv[1:])
